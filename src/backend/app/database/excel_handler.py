"""
Excel/Spreadsheet Handler
Manages Excel, CSV, and ODS files
"""

from typing import Dict, List, Optional, Any
from pathlib import Path
import json


class ExcelHandler:
    """Handle Excel and spreadsheet operations"""
    
    def __init__(self, base_path: Path):
        self.base_path = Path(base_path)
        self.base_path.mkdir(parents=True, exist_ok=True)
    
    def get_sheet_names(self, file_path: str) -> List[str]:
        """Get list of sheet names in Excel file"""
        try:
            # Try using openpyxl for .xlsx files
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets
        except ImportError:
            return ["Sheet1"]
        except Exception as e:
            return [f"Error: {str(e)}"]
    
    def read_sheet(self, file_path: str, sheet_name: Optional[str] = None) -> Dict:
        """Read data from Excel sheet"""
        try:
            
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Use first sheet if not specified
            if sheet_name is None:
                sheet_name = wb.sheetnames[0]
            
            sheet = wb[sheet_name]
            
            # Extract data
            data = []
            headers = []
            
            for row_idx, row in enumerate(sheet.iter_rows()):
                row_data = [cell.value for cell in row]
                
                if row_idx == 0:
                    headers = row_data
                else:
                    data.append(dict(zip(headers, row_data)))
            
            wb.close()
            
            return {
                "file": file_path,
                "sheet": sheet_name,
                "headers": headers,
                "data": data,
                "row_count": len(data),
                "column_count": len(headers)
            }
            
        except ImportError:
            return {"error": "openpyxl not installed. Install with: pip install openpyxl"}
        except Exception as e:
            return {"error": str(e)}
    
    def read_csv(self, file_path: str) -> Dict:
        """Read CSV file"""
        import csv
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                data = list(reader)
                
                return {
                    "file": file_path,
                    "headers": list(data[0].keys()) if data else [],
                    "data": data,
                    "row_count": len(data)
                }
        except Exception as e:
            return {"error": str(e)}
    
    def get_preview(self, file_path: str, rows: int = 10) -> Dict:
        """Get preview of spreadsheet (first N rows)"""
        ext = Path(file_path).suffix.lower()
        
        if ext == '.csv':
            result = self.read_csv(file_path)
        else:
            result = self.read_sheet(file_path)
        
        if "error" in result:
            return result
        
        # Limit rows for preview
        result["data"] = result["data"][:rows]
        result["preview_rows"] = rows
        
        return result
    
    def create_excel(self, data: List[Dict], output_path: str) -> str:
        """Create new Excel file from data"""
        try:
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            if not data:
                wb.save(output_path)
                return output_path
            
            # Get headers from first row
            headers = list(data[0].keys())
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(output_path)
            return output_path
            
        except ImportError:
            # Fallback to CSV
            csv_path = output_path.replace('.xlsx', '.csv')
            self.create_csv(data, csv_path)
            return csv_path
        except Exception as e:
            return f"Error: {str(e)}"
    
    def create_csv(self, data: List[Dict], output_path: str) -> str:
        """Create CSV file from data"""
        
        if not data:
            return output_path
        
        headers = list(data[0].keys())
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(data)
        
        return output_path
    
    def export_sheet_to_csv(self, file_path: str, sheet_name: str, output_path: str) -> str:
        """Export Excel sheet to CSV"""
        data = self.read_sheet(file_path, sheet_name)
        
        if "error" in data:
            return data["error"]
        
        return self.create_csv(data["data"], output_path)
    
    def merge_excel_files(self, file_paths: List[str], output_path: str) -> str:
        """Merge multiple Excel files into one"""
        try:
            
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            for file_path in file_paths:
                try:
                    source_wb = openpyxl.load_workbook(file_path, data_only=True)
                    sheet_name = Path(file_path).stem[:31]  # Excel sheet name limit
                    
                    # Copy first sheet
                    source_sheet = source_wb[source_wb.sheetnames[0]]
                    target_sheet = wb.create_sheet(title=sheet_name)
                    
                    for row in source_sheet.iter_rows():
                        target_sheet.append([cell.value for cell in row])
                    
                    source_wb.close()
                except Exception as e:
                    print(f"Error processing {file_path}: {e}")
            
            wb.save(output_path)
            return output_path
            
        except ImportError:
            return "openpyxl not installed"
        except Exception as e:
            return f"Error: {str(e)}"
    
    def query_data(self, file_path: str, filters: Dict[str, Any]) -> List[Dict]:
        """Query data with filters"""
        ext = Path(file_path).suffix.lower()
        
        if ext == '.csv':
            result = self.read_csv(file_path)
        else:
            result = self.read_sheet(file_path)
        
        if "error" in result:
            return result
        
        data = result["data"]
        
        # Apply filters
        filtered = []
        for row in data:
            match = True
            for key, value in filters.items():
                if key in row and row[key] != value:
                    match = False
                    break
            if match:
                filtered.append(row)
        
        return filtered
